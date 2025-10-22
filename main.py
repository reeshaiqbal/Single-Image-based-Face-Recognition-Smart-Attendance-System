# -*- coding: utf-8 -*-
import os
import zipfile
from tkinter import *
from tkinter import ttk
from tkinter.ttk import Treeview
import datetime
import openpyxl
import pandas as pd
from openpyxl import *
import subprocess
import csv
from PIL import Image, ImageTk
import datetime
import cv2
import PIL.Image
if not hasattr(PIL.Image, 'Resampling'):
    # Pillow<9.0
    PIL.Image.Resampling = PIL.Image
from keras.models import load_model
from keras.models import Model
import tensorflow as tf
import pandas as pd
from datetime import datetime
from openpyxl import *
import os
import cv2
import numpy as np
from os import listdir
from os.path import isfile, join
from retinaface import RetinaFace
from tkinter import filedialog as fd
from tkfilebrowser import askopendirname
from tkinter import messagebox
from keras.layers import Input, Lambda, Dense, Flatten
from keras.models import Model
from keras.applications.xception import Xception
from keras.applications.xception import preprocess_input
from keras.preprocessing import image
from keras.preprocessing.image import ImageDataGenerator
from keras.models import Sequential
import numpy as np
from glob import glob
import matplotlib.pyplot as plt

class Home_Page:
    def __init__(self, main_window):
        self.img_name = None
        self.L2 = None
        self.btn7 = None
        self.btn6 = None
        self.frame = None
        self.L1 = None
        self.cam = None
        self.btn5 = None
        self.btn8 = None
        self.f3 = None
        self.f4 = None
        self.L3 = None
        self.btn9 = None
        self.btn10 = None
        self.options = None
        self.dropdown = None
        self.btn12 = None
        self.btn13 = None
        self.btn14 = None
        self.Trfolder = None
        self.Tsfolder = None
        self.faces = []
        self.AttendanceReportTable = None
        self.L4 = None
        self.L5 = None
        self.L6 = None
        self.L7 = None
        self.L8 = None
        self.L9 = None
        self.Record = []
        self.win = main_window
        self.f1 = Frame(self.win, bg='#000000').place(height=150, width=957, x=0, y=0)
        self.pil_img = Image.open("/home/reesh/title_image.png")
        self.img = ImageTk.PhotoImage(self.pil_img.resize((957, 150), Image.Resampling.LANCZOS))
        self.lbl_img = Label(self.f1, text="Face Recognition Attendance System", fg="#fff",
                             font=('Times New Roman', 17, 'bold'),
                             image=self.img,
                             compound='center')
        self.lbl_img.img = self.img
        self.lbl_img.place(height=150, width=757, x=200, y=0)

        self.pil_img1 = Image.open("/home/reesh/logo1.png")
        self.img1 = ImageTk.PhotoImage(self.pil_img1.resize((150, 150), Image.Resampling.LANCZOS))
        self.lbl_img1 = Label(self.f1, bg='#FFFFFA',
                              image=self.img1)
        self.lbl_img1.img1 = self.img1
        self.lbl_img1.place(height=150, width=200, x=0, y=0)

        self.f2 = Frame(self.win, bg="#6B6B8E").place(height=544, width=200, x=0, y=150)

        self.btn1 = Button(self.f2, text="Start Attendance", relief="ridge", borderwidth="5",
                           command=self.Start_Attendance)
        self.btn1.place(height=30, width=150, x=25, y=200)

        self.btn2 = Button(self.f2, text="View Record", command=self.view_Record, relief="ridge", borderwidth="5")
        self.btn2.place(height=30, width=150, x=25, y=300)

        self.btn3 = Button(self.f2, text="Train new dataset", command = self.SelectFolder, relief="ridge", borderwidth="5")
        self.btn3.place(height=30, width=150, x=25, y=400)

        self.btn4 = Button(self.f2, text="Exit", command = self.destroy_window, relief="ridge", borderwidth="5")
        self.btn4.place(height=30, width=150, x=25, y=500)

        self.f3 = Frame(self.win, bg='#CDCDC8', highlightbackground="#45458B", highlightthickness=10, )
        self.f3.place(height=394, width=757, x=200, y=150)

        self.frame_pixel = None
        self.counter = 0
        self.temp = True
        self.check_submit = False

    ###################################Statrt Atendance########################################################

    def Start_Attendance(self):
        if(self.f4 is not None):
            self.f4.destroy()
            self.f3 = Frame(self.win, bg='#CDCDC8', highlightbackground="#45458B", highlightthickness=10, )
            self.f3.place(height=394, width=757, x=200, y=150)
            

        self.cam = cv2.VideoCapture("http://10.44.195.200:8080/video")
        self.btn5 = Button(self.f3, text="SnapShot", relief="ridge", command=self.snapcapture, borderwidth="5",
                           highlightcolor="white")
        self.btn5.place(height=30, width=737, x=0, y=345)
        self.L1 = Label(self.f3, bg="black", borderwidth=10, highlightcolor="white")
        while True:
            if (self.temp == False):
                break
            ret, self.frame = self.cam.read()
            self.frame = cv2.resize(self.frame, (718, 325), )
            self.frame_pixel = self.frame
            print(self.frame)
            self.frame = ImageTk.PhotoImage(Image.fromarray(self.frame))
            self.L1.config(image=self.frame)
            self.L1.pack()
            self.f3.update()
        self.cam.release()
        cv2.destroyAllWindows()

    def snapcapture(self):
        self.counter += 1
        print(type(self.frame_pixel))
        print("Photo saved")
        try:
            self.L1.configure(image=self.frame)
            self.L1.image = self.frame
            self.temp = False
            self.btn5.destroy()
            self.btn6 = Button(self.f3, text="Ok", command=self.save_image, relief="ridge", borderwidth="5",
                               highlightcolor="white")
            self.btn6.place(height=30, width=367, x=1, y=345)

            self.btn7 = Button(self.f3, text="Cancel", command=self.retry_cam, relief="ridge", borderwidth="5",
                               highlightcolor="white")
            self.btn7.place(height=30, width=367, x=367, y=345)
        except:
            print("error in snap calling function")

    def save_image(self):
        self.frame = ImageTk.PhotoImage(Image.fromarray(self.frame_pixel))
        time = str(datetime.now().today()).replace(":", '') + ".jpg"
        print(time)
        i = 0
        self.img_name = f"{time} image number{self.counter}.jpg"
        cv2.imwrite(self.img_name, self.frame_pixel)
        print("{} written!".format(self.img_name))
        self.Generate_attandance_sheet()

    def retry_cam(self):
        self.btn6.destroy()
        self.btn7.destroy()
        self.L1.destroy()
        if (self.check_submit == True):
            self.btn8.destroy()

        self.temp = True
        self.Start_Attendance()
        
    def destroy_window(self):
        self.win.destroy()

    def Generate_attandance_sheet(self):
        
        
        
        ###############################################################################################
        classifier = load_model(filepath = r'/home/reesh/.config/spyder-py3/classDataset/Custom_Train_Model/Xception_new_model.h5')
        monkey_breeds_dict = {"[0]": "Arshad_Waseer(sp19-bcs-087)", 
                      "[1]": "Usman_Khalid(sp19-bcs-079)",
                      "[2]": "Nisar_Ahmad(sp19-bcs-107)",
                      "[3]": "Shahzad_Iqbal(sp19-bcs-097)",
                      "[4]": "Zaffar_Farooq(sp19-bcs-088)"} 
        def draw_test(name, pred, im):
            monkey = monkey_breeds_dict[str(pred)]
            BLACK = [0,0,0]
            expanded_image = cv2.copyMakeBorder(im, 80, 0, 0, 100 ,cv2.BORDER_CONSTANT,value=BLACK)
            cv2.putText(expanded_image, monkey, (20, 60) , cv2.FONT_HERSHEY_SIMPLEX,1, (0,0,255), 2)
            cv2.imshow(name, expanded_image)
            cv2.waitKey(0)
        self.faces = RetinaFace.extract_faces(img_path = "/home/reesh/Pictures/Screenshot from 2023-01-01 12-42-07.png", align = True)
        
        for i, face in enumerate(self.faces):
            input_original = face.copy()
            input_original = cv2.resize(input_original, None, fx=0.5, fy=0.5, interpolation = cv2.INTER_LINEAR)
            face = cv2.resize(face, (299, 299), interpolation = cv2.INTER_LINEAR)
            face = face / 255.
            face = face.reshape(1,299,299,3) 
            res = np.argmax(classifier.predict(face, 1, verbose = 0), axis=1)
            predd = monkey_breeds_dict[str(res)]
            preddlist = predd.split("(")
            Name = preddlist[0]
            Roll_no = preddlist[-1].strip(")")
            print("Name = ", Name, "\n"+"Roll_no = ", Roll_no)
            datetime_obj = datetime.now()
            tm = datetime_obj.time()
            row = [f"{i+1}", f"{Name}", f"{Roll_no}",  "BSCS", f"{tm}", "present"]
            self.Record.append(row)
            
        
        ################################################################################################
        self.L1.destroy()
        self.btn6.destroy()
        self.btn8 = Button(self.f3, text="Submit_sheet", command=self.submit_sheet, relief="ridge", borderwidth="5",
                           highlightcolor="white")
        self.btn8.place(height=30, width=367, x=1, y=345)
        self.check_submit = True
        self.L2 = Label(self.f3, bg="black",  highlightcolor="white")
        scroll_x = Scrollbar(self.L2, orient=HORIZONTAL)
        scroll_y = Scrollbar(self.L2, orient=VERTICAL)
        s = ttk.Style()
        s.theme_use('clam')
        s.map("Custom.Treeview", background=[("selected", "green")])
        s.configure('Treeview')
        self.AttendanceReportTable = Treeview(self.L2,
                                         columns=("id", "roll", "name", "department", "time", "Attendance Status"),
                                         height=20, xscrollcommand=scroll_x.set, yscrollcommand=scroll_y.set)
        scroll_x.pack(side=BOTTOM, fill=X)
        scroll_y.pack(side=RIGHT, fill=Y)
        scroll_x.config(command=self.AttendanceReportTable.xview)
        scroll_y.config(command=self.AttendanceReportTable.yview)
        self.AttendanceReportTable.heading("id", text="Attendance ID")
        self.AttendanceReportTable.heading("roll", text=" Roll")
        self.AttendanceReportTable.heading("name", text="Name")
        self.AttendanceReportTable.heading("time", text="Time")
        self.AttendanceReportTable.heading("department", text="Department")

        self.AttendanceReportTable.heading("Attendance Status", text="Attendance Status")
        self.AttendanceReportTable["show"] = "headings"
        self.AttendanceReportTable.column("id", width=100)
        self.AttendanceReportTable.column("roll", width=100)
        self.AttendanceReportTable.column("name", width=120)
        self.AttendanceReportTable.column("department", width=100)
        self.AttendanceReportTable.column("time", width=100)
        self.AttendanceReportTable.column("Attendance Status", width=100)
        self.AttendanceReportTable.place(height=330, width=720, x=0, y=0)
        
        for i in range(len(self.faces)):
            self.AttendanceReportTable.insert("", END, values = self.Record[i])
        self.L2.place(height=345, width=738, x=0, y=0)

    def submit_sheet(self):
        datetime_obj = datetime.now()
        dt = datetime_obj.date()
        tm = datetime_obj.time()
        df = pd.DataFrame(self.Record,
        columns=['Student ID', 'Student Name','Registartion Number', 'Department' , 'Time','Attendance Status'])
        book = load_workbook(r'/home/reesh/Attendance_Record.xlsx')
        with pd.ExcelWriter('/home/reesh/Attendance_Record.xlsx', engine='openpyxl') as writer:
                    if book is not None:
                        writer.book = book
                    df.to_excel(writer, str(dt), index=False)
                    
    def view_Record(self):
        if(self.f3 is not None):
            self.f3.destroy()
        
        self.f4 = Frame(self.win, bg='#CDCDC8', highlightbackground="#45458B", highlightthickness=10, )
        self.f4.place(height=394, width=757, x=200, y=150)
        
        self.L3 = Label(self.f4, bg="black", borderwidth=10, highlightcolor="white")
        
        wb = load_workbook(r'/home/reesh/Attendance_Record.xlsx')
        dates = wb.sheetnames
        
        self.options = StringVar(self.f4)
        self.options.set("Select Date")
        # self.dropdown = tk.OptionMenu(self.f4, self.options, *dates, self.view_sheets)
        self.dropdown = OptionMenu(self.f4, self.options, *dates, command = self.view_sheets)
        self.dropdown.place(width=115, x= 0, y = 0)
        self.dropdown.config(bg = "#45458B", fg = "white")
        
        scroll_x = Scrollbar(self.L3, orient=HORIZONTAL)
        scroll_y = Scrollbar(self.L3, orient=VERTICAL)
        s = ttk.Style()
        s.theme_use('clam')
        s.map("Custom.Treeview", background=[("selected", "green")])
        s.configure('Treeview')
        self.AttendanceReportTable = Treeview(self.L3,
                                         columns=("id", "roll", "name", "department", "time", "Att_Status"),
                                         height=20, xscrollcommand=scroll_x.set, yscrollcommand=scroll_y.set)
        scroll_x.pack(side=BOTTOM, fill=X)
        scroll_y.pack(side=RIGHT, fill=Y)
        scroll_x.config(command=self.AttendanceReportTable.xview)
        scroll_y.config(command=self.AttendanceReportTable.yview)
        self.AttendanceReportTable.heading("id", text="Attendance ID")
        self.AttendanceReportTable.heading("roll", text=" Roll")
        self.AttendanceReportTable.heading("name", text="Name")
        self.AttendanceReportTable.heading("time", text="Time")
        self.AttendanceReportTable.heading("department", text="Department")

        self.AttendanceReportTable.heading("Att_Status", text="Att_Status")
        self.AttendanceReportTable["show"] = "headings"
        self.AttendanceReportTable.column("id", width=100)
        self.AttendanceReportTable.column("roll", width=100)
        self.AttendanceReportTable.column("name", width=120)
        self.AttendanceReportTable.column("department", width=100)
        self.AttendanceReportTable.column("time", width=100)
        self.AttendanceReportTable.column("Att_Status", width=100)
        self.AttendanceReportTable.pack()
        for i in range(5):
            self.AttendanceReportTable.insert("", END, values = [f'{i}', 'Yusra Qayyum', 'sp19-bcs-096', 'bcs', '2:00', 'Ghaib'])
       
        self.L3.place(height=375, width=620, x=117, y=0)
        
    def view_sheets(self, options):
        
        sheetn = self.options.get()
        
        df1 = pd.read_excel(r'/home/reesh/Attendance_Record.xlsx', sheetn)
        
        for item in self.AttendanceReportTable.get_children():
            self.AttendanceReportTable.delete(item)
        for index, row in df1.iterrows():
            
            self.AttendanceReportTable.insert("", END, values = [row['Student ID'], row['Student Name'], row['Registartion Number'], row['Department'], row['Time'], row['Attendance Status']])
            print(row['Student ID'], row['Student Name'], row['Registartion Number'], row['Department'], row['Time'], row['Attendance Status'])
    
    def SelectFolder(self):
        Label(self.f3,  text= "Selected training direactory:").place(x=2, y=100)
        self.btn12 = Button(self.f3, text="Select training directory", command=self.SelecttrDirectory, relief="ridge", borderwidth="5",
                           highlightcolor="white")
        self.btn12.place(height=30, width=245, x=1, y=345)
        Label(self.f3,  text= "Selected testing direactory:").place(x=246, y=100)
        self.btn13 = Button(self.f3, text="Select testing directory", command=self.SelecttsDirectory, relief="ridge", borderwidth="5",
                           highlightcolor="white")
        self.btn13.place(height=30, width=245, x=246, y=345)
        
        self.btn14 = Button(self.f3, text="Start traning", command=self.Start_training, relief="ridge", borderwidth="5",
                           highlightcolor="white")
        self.btn14.place(height=30, width=245, x=491, y=345)
        
    
    def SelecttrDirectory(self):
        self.Trfolder = askopendirname(parent=self.f3, title = "Select folder")
        Label(self.f3,  text= self.Trfolder).place(x=2, y=120, width=245)
        print(self.Trfolder)
    
    def SelecttsDirectory(self):
        self.Tsfolder = askopendirname(parent=self.f3, title = "Select folder")
        Label(self.f3,  text= self.Tsfolder).place(x=246, y=120, width=245)
        print(self.Tsfolder)
    def Start_training(self):
        Label(self.f3,  text= "Start training .....").place(x=491, y=100)
        if(self.Trfolder is not None and self.Tsfolder is not None):
            IMAGE_SIZE = [299, 299]
            train_path = f'{self.Trfolder}/'.replace("/", "//")
            valid_path = f'{self.Tsfolder}/'.replace("/", "//")
            X = Xception(input_shape=IMAGE_SIZE + [3], weights='imagenet', include_top=False)
            for layer in X.layers:
              layer.trainable = False
            folders = glob('/home/reesh/.config/spyder-py3/classDataset/TrainData/*')
            x = Flatten()(X.output)
            prediction = Dense(len(folders), activation='softmax')(x)
            model = Model(inputs=X.input, outputs=prediction)
            model.summary()
            model.compile(
              loss='categorical_crossentropy',
              optimizer='adam',
              metrics=['accuracy']
            )
            
            
            from keras.preprocessing.image import ImageDataGenerator
            
            train_datagen = ImageDataGenerator(rescale = 1./255,
                                               shear_range = 0.2,
                                               zoom_range = 0.2,
                                               horizontal_flip = True)
            
            test_datagen = ImageDataGenerator(rescale = 1./255)
            
            training_set = train_datagen.flow_from_directory(train_path,
                                                             target_size = (299, 299),
                                                             batch_size = 32,
                                                             class_mode = 'categorical')
            
            test_set = test_datagen.flow_from_directory(valid_path,
                                                        target_size = (299, 299),
                                                        batch_size = 32,
                                                        class_mode = 'categorical')
            
            '''r=model.fit_generator(training_set,
                                     samples_per_epoch = 8000,
                                     nb_epoch = 5,
                                     validation_data = test_set,
                                     nb_val_samples = 2000)'''
            
           
            r = model.fit_generator(
              training_set,
              validation_data=test_set,
              epochs=15,
              steps_per_epoch=len(training_set),
              validation_steps=len(test_set)
            )
            # loss
            plt.plot(r.history['loss'], label='train loss')
            plt.plot(r.history['val_loss'], label='val loss')
            plt.legend()
            plt.show()
            plt.savefig('LossVal_loss')
            
            # accuracies
            plt.plot(r.history['accuracy'], label='train acc')
            plt.plot(r.history['val_accuracy'], label='val acc')
            plt.legend()
            plt.show()
            plt.savefig('AccVal_acc')
            
            model.save(r'/home/reesh/.config/spyder-py3/classDataset/VGG16rained-model/Xception_new_model.h5')
        else:
            messagebox.showinfo("Sorry","Please Select train test directory")

if __name__ == "__main__":
    os.system("/home/reesh/Desktop/Single Image based face recognition by two-fold NN Model for Smart Attendance/Splash_Screen.py")
    process1 = subprocess.Popen(["python", r"/home/reesh/Desktop/Single Image based face recognition by two-fold NN Model for Smart Attendance/Splash_Screen.py"])
    process1.wait()
    window = Tk()
    window.geometry("957x544+200+0")

    obj = Home_Page(window)
    window.mainloop()
